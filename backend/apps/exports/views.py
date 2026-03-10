"""
Export views: XLSX results and PDF evaluation reports.
"""
import io
import re
from decimal import ROUND_HALF_UP, Decimal
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import permissions
from rest_framework.views import APIView

from apps.evaluations.services import calculate_final_grade
from apps.exams.models import Exam
from apps.evaluations.models import Evaluation
from apps.exams.models import StationAssignment
from apps.users.models import User


def _safe_filename(value: str) -> str:
    """Strip characters that could break Content-Disposition headers."""
    return re.sub(r'[^\w\-.]', '_', value)


class ExamResultsXlsxView(APIView):
    """GET /exams/{exam_id}/exports/results.xlsx"""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, exam_id):
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        exam = get_object_or_404(Exam, pk=exam_id)

        # RBAC: evaluators can only export exams where they have assignments
        if request.user.role == User.Role.EVALUATOR:
            if not StationAssignment.objects.filter(exam=exam, evaluator=request.user).exists():
                return HttpResponse(status=403)

        active_stations = list(
            exam.stations.filter(is_active=True).order_by("order", "id")
        )
        results = calculate_final_grade(exam)

        wb = openpyxl.Workbook()

        # ── Sheet: Resultados ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "Resultados"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E3A8A")
        center = Alignment(horizontal="center", vertical="center")

        base_headers = ["RUT", "Nombre", "Correo"]
        station_headers = [
            f"Nota Estación {s.order} - {s.name}" for s in active_stations
        ]
        all_headers = base_headers + station_headers + ["Nota Final", "Aprobado"]

        for col_idx, header in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Data rows
        for row_idx, r in enumerate(results, start=2):
            student = r["student"]
            ws.cell(row=row_idx, column=1, value=student.rut)
            ws.cell(row=row_idx, column=2, value=student.full_name)
            ws.cell(row=row_idx, column=3, value=student.email)

            for st_idx, station in enumerate(active_stations):
                grade = r["station_grades"].get(station.id)
                val = float(grade) if grade is not None else None
                ws.cell(row=row_idx, column=4 + st_idx, value=val)
                if val is not None:
                    ws.cell(row=row_idx, column=4 + st_idx).number_format = "0.00"

            final_col = 4 + len(active_stations)
            approved_col = final_col + 1

            if r["final_grade"] is not None:
                ws.cell(row=row_idx, column=final_col, value=float(r["final_grade"]))
                ws.cell(row=row_idx, column=final_col).number_format = "0.00"
                ws.cell(
                    row=row_idx,
                    column=approved_col,
                    value="Aprobado" if r["approved"] else "Reprobado",
                )
            else:
                ws.cell(row=row_idx, column=final_col, value=None)
                ws.cell(row=row_idx, column=approved_col, value="Sin datos")

        # Freeze pane + autofilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Column widths
        column_widths = {"A": 14, "B": 30, "C": 28}
        for col_idx in range(len(all_headers)):
            col_letter = get_column_letter(col_idx + 1)
            if col_letter not in column_widths:
                column_widths[col_letter] = 20
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ── Sheet: Metadata ────────────────────────────────────────────────
        ws_meta = wb.create_sheet("Metadata")
        ws_meta.column_dimensions["A"].width = 25
        ws_meta.column_dimensions["B"].width = 40

        meta_rows = [
            ("ECOE", exam.name),
            ("Descripción", exam.description),
            ("Estado", exam.get_status_display()),
            (
                "Fecha inicio",
                str(exam.start_date) if exam.start_date else "-",
            ),
            ("Fecha exportación", timezone.now().strftime("%d/%m/%Y %H:%M")),
            ("", ""),
            ("Ponderaciones por estación", ""),
        ]
        for row in meta_rows:
            ws_meta.append(row)

        for station in active_stations:
            ws_meta.append(
                [
                    f"Estación {station.order}: {station.name}",
                    f"{station.weight_percent}%",
                ]
            )

        # Respond
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"resultados_ecoe_{exam.id}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class EvaluationPdfView(APIView):
    """GET /evaluations/{evaluation_id}/exports/evaluation.pdf"""

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, evaluation_id):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
            HRFlowable,
        )

        evaluation = get_object_or_404(
            Evaluation.objects.select_related(
                "exam", "station", "student", "evaluator"
            ).prefetch_related("item_scores__rubric_item"),
            pk=evaluation_id,
        )

        # Permission check
        from apps.users.models import User
        if request.user.role == User.Role.EVALUATOR:
            if evaluation.evaluator != request.user and not StationAssignment.objects.filter(
                exam=evaluation.exam,
                station=evaluation.station,
                evaluator=request.user,
            ).exists():
                from rest_framework.response import Response
                from rest_framework import status as drf_status
                return HttpResponse(status=403)

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=4,
            textColor=colors.HexColor("#1E3A8A"),
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent=styles["Normal"],
            fontSize=10,
            spaceAfter=2,
        )
        small_style = ParagraphStyle(
            "Small",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
        )
        normal_style = styles["Normal"]
        normal_style.fontSize = 9

        story = []

        # ── Header ────────────────────────────────────────────────────────
        story.append(Paragraph("Pauta de Evaluación ECOE", title_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1E3A8A")))
        story.append(Spacer(1, 0.3 * cm))

        finalized_str = (
            evaluation.finalized_at.strftime("%d/%m/%Y %H:%M")
            if evaluation.finalized_at
            else "No finalizado"
        )

        header_data = [
            ["ECOE:", evaluation.exam.name, "Fecha:", finalized_str],
            ["Estación:", evaluation.station.name, "Estado:", evaluation.get_status_display()],
            ["Estudiante:", evaluation.student.full_name, "RUT:", evaluation.student.rut],
            ["Evaluador:", evaluation.evaluator.full_name, "Ponderación:", f"{evaluation.station.weight_percent}%"],
        ]
        header_table = Table(header_data, colWidths=[3 * cm, 7 * cm, 3 * cm, 4 * cm])
        header_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(header_table)
        story.append(Spacer(1, 0.5 * cm))

        # ── Rubric Table ───────────────────────────────────────────────────
        story.append(Paragraph("Detalle de Ítems Evaluados", ParagraphStyle(
            "SectionTitle", parent=styles["Heading2"], fontSize=11,
            textColor=colors.HexColor("#1E3A8A"), spaceAfter=4,
        )))

        table_data = [["N°", "Descripción del Ítem", "Pts. Máx.", "Pts. Obtenidos", "Observación"]]
        item_scores = sorted(
            evaluation.item_scores.all(),
            key=lambda s: (s.rubric_item.order, s.rubric_item.id),
        )

        for idx, score in enumerate(item_scores, start=1):
            points_str = (
                str(score.points.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                if score.points is not None
                else "-"
            )
            max_str = str(
                score.rubric_item.max_points.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            )
            table_data.append(
                [
                    str(idx),
                    Paragraph(score.rubric_item.description, ParagraphStyle("Cell", fontSize=8, leading=10)),
                    max_str,
                    points_str,
                    Paragraph(score.comment or "", ParagraphStyle("Cell", fontSize=8, leading=10)),
                ]
            )

        rubric_table = Table(
            table_data,
            colWidths=[1 * cm, 7.5 * cm, 2 * cm, 2.5 * cm, 4 * cm],
            repeatRows=1,
        )
        rubric_table.setStyle(
            TableStyle(
                [
                    # Header row
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    # Data rows
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("ALIGN", (0, 1), (0, -1), "CENTER"),
                    ("ALIGN", (2, 1), (3, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(rubric_table)
        story.append(Spacer(1, 0.5 * cm))

        # ── Totals ─────────────────────────────────────────────────────────
        total_str = (
            str(evaluation.total_points.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            if evaluation.total_points is not None
            else "-"
        )
        grade_str = (
            str(evaluation.grade.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            if evaluation.grade is not None
            else "-"
        )
        approved = None
        if evaluation.grade is not None:
            approved = evaluation.grade >= Decimal("4.0")
        approved_str = "APROBADO" if approved else ("REPROBADO" if approved is False else "-")
        approved_color = (
            colors.HexColor("#16A34A") if approved
            else (colors.HexColor("#DC2626") if approved is False else colors.grey)
        )

        totals_data = [
            ["Puntaje Total:", total_str, "Nota Estación:", grade_str, "Resultado:", approved_str],
        ]
        totals_table = Table(
            totals_data,
            colWidths=[3.5 * cm, 2.5 * cm, 3 * cm, 2.5 * cm, 2.5 * cm, 3 * cm],
        )
        totals_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("ALIGN", (1, 0), (1, 0), "CENTER"),
                    ("ALIGN", (3, 0), (3, 0), "CENTER"),
                    ("ALIGN", (5, 0), (5, 0), "CENTER"),
                    ("TEXTCOLOR", (5, 0), (5, 0), approved_color),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFDBFE")),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(totals_table)
        story.append(Spacer(1, 0.5 * cm))

        # ── General Comment ────────────────────────────────────────────────
        story.append(Paragraph("Observación General", ParagraphStyle(
            "SectionTitle", parent=styles["Heading2"], fontSize=11,
            textColor=colors.HexColor("#1E3A8A"), spaceAfter=4,
        )))

        comment_text = evaluation.general_comment or "(Sin observación general)"
        comment_box = Table(
            [[Paragraph(comment_text, ParagraphStyle("Comment", fontSize=9, leading=12))]],
            colWidths=[17 * cm],
        )
        comment_box.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FAFAFA")),
                ]
            )
        )
        story.append(comment_box)
        story.append(Spacer(1, 1.5 * cm))

        # ── Signature ─────────────────────────────────────────────────────
        sig_data = [
            [
                "_" * 35 + "",
                "",
                "_" * 35 + "",
            ],
            [
                "Firma Evaluador",
                "",
                "Fecha",
            ],
            [
                evaluation.evaluator.full_name,
                "",
                finalized_str,
            ],
        ]
        sig_table = Table(sig_data, colWidths=[7 * cm, 3 * cm, 7 * cm])
        sig_table.setStyle(
            TableStyle(
                [
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (2, 0), (2, -1), "CENTER"),
                    ("TEXTCOLOR", (0, 1), (-1, 1), colors.grey),
                    ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
                ]
            )
        )
        story.append(sig_table)

        # ── Footer note ────────────────────────────────────────────────────
        story.append(Spacer(1, 0.5 * cm))
        story.append(
            Paragraph(
                f"Documento generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} · ECOE MVP",
                small_style,
            )
        )

        doc.build(story)
        buffer.seek(0)

        student_name = _safe_filename(evaluation.student.full_name.replace(" ", "_"))
        station_name = _safe_filename(evaluation.station.name.replace(" ", "_"))
        filename = f"evaluacion_{student_name}_{station_name}_{evaluation.id}.pdf"
        response = HttpResponse(buffer.read(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
